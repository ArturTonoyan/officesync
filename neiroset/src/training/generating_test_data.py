import random
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment
import os

# Возможные типы оборудования
typeEquipment = [
    {"id": 1, "name": "Принтер"},
    {"id": 2, "name": "Компьютер"},
    {"id": 3, "name": "Кондиционер"},
    {"id": 4, "name": "Стол"},
]

# Возможные типы отказов
typeFaults = {
    0: "Оборудование исправно",
    1: "Скрытый отказ",
    2: "Неполный отказ",
    3: "Критический отказ",
}

def random_date(start_year=2020, end_year=2025):
    """Генерирует случайную дату в диапазоне лет"""
    start = datetime(start_year, 1, 1)
    end = datetime(end_year, 12, 31)
    random_days = random.randint(0, (end - start).days)
    return (start + timedelta(days=random_days))

def calculate_wear(work_hours, max_work_hours, previous_failures, failure_frequency, operating_conditions):
    """Вычисляет износ с учетом наработанных часов, отказов, частоты отказов и условий эксплуатации"""
    
    # Расчет базового износа на основе наработанных часов
    base_wear = (work_hours / max_work_hours) * 0.9
    
    # Дополнительный износ из-за частоты отказов (например, количество отказов на 1000 часов работы)
    failure_wear = failure_frequency * 0.2
    
    # Дополнительный износ из-за условий эксплуатации (например, экстремальные условия)
    condition_wear = operating_conditions * 0.15  # Множитель для условий эксплуатации (например, 0.1 для обычных условий, 0.2 для сложных)

    # Износ на основе отказов (критические и неполные)
    failure_impact = 0
    for failure in previous_failures:
        if failure["type_fault"] == 3:  # Критический отказ
            failure_impact += 0.3
        elif failure["type_fault"] == 2:  # Неполный отказ
            failure_impact += 0.15
    
    # Суммируем все элементы
    total_wear = base_wear + failure_wear + condition_wear + failure_impact
    
    # Ограничиваем максимальный износ до 1.0
    return min(total_wear, 1.0)

def generate_failure():
    """Генерирует случайный отказ оборудования"""
    date_start = random_date()
    # Генерация даты окончания отказа, которая должна быть позже начала
    date_end = date_start + timedelta(days=random.randint(1, 10))
    return {
        "id": f"to{random.randint(100, 999)}",
        "date_start": date_start.date(),
        "date_end": date_end.date(),
        "type_fault": random.choice(list(typeFaults.keys())),
        "cost_to": random.randint(500, 5000),
    }

def generate_equipment_case():
    """Генерирует случайный кейс оборудования с учетом закономерностей"""
    equipment_type = random.choice(typeEquipment)
    
    # Генерируем начальную наработку и максимальную наработку
    max_work_hours = random.choice([5000, 8000, 10000, 12000])
    work_hours = random.randint(0, max_work_hours)

    # Периодичность ТО
    maintenance_frequency = random.choice([30, 90, 180, 365])

    # Генерируем случайный процент использования оборудования
    average_daily_usage_hours = round(random.uniform(1.0, 8.0), 1)

    # Генерируем случайные отказы
    previous_failures = [generate_failure() for _ in range(random.randint(0, 5))]

    # Вычисляем износ
    wear = calculate_wear(work_hours, max_work_hours, previous_failures, failure_frequency=0.1, operating_conditions=0.1)

    case = {
        "id": str(random.randint(10000, 99999)),
        "type": equipment_type["id"],
        "type_name": equipment_type["name"],
        "date_start": random_date().date(),
        "work_hours": work_hours,
        "max_work_hours": max_work_hours,
        "average_daily_usage_hours": average_daily_usage_hours,
        "maintenance_frequency_days": maintenance_frequency,
        "previous_failures_types": previous_failures,
        "wear": wear,
    }
    return case

def clean_data(cases):
    """Очищает данные от некорректных или ненужных записей"""
    cleaned_cases = []
    for case in cases:
        # Убираем записи с нулевой наработкой или отрицательным износом
        if case["work_hours"] == 0 or case["wear"] <= 0:
            continue
        
        # Убираем оборудование с ненадежными датами
        if case["date_start"] > datetime.today().date():
            continue
        
        # Применяем другие условия фильтрации по необходимости
        if case['wear'] >= 1:
            continue
        
        cleaned_cases.append(case)
    return cleaned_cases

def save_cases_to_excel(cases, filename="test_cases.xlsx"):
    """Сохраняет список кейсов в Excel-файл с настройкой ширины столбцов и выравниванием текста"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Equipment Cases"

    # Заголовки столбцов
    headers = [
        "id", "type", "type_name", "date_start",
        "work_hours", "max_work_hours", "average_daily_usage_hours",
        "maintenance_frequency_days", "wear", "previous_failures"
    ]

    ws.append(headers)

    # Заполнение данных
    for case in cases:
        previous_failures = "; ".join([ 
            f"ID: {failure['id']}, Тип: {typeFaults[failure['type_fault']]}, "
            f"Дата начала: {failure['date_start']}, Дата окончания: {failure['date_end']}, "
            f"Стоимость: {failure['cost_to']} "
            for failure in case["previous_failures_types"]
        ])
        ws.append([
            case["id"],
            case["type"],
            case["type_name"],
            case["date_start"],
            case["work_hours"],
            case["max_work_hours"],
            case["average_daily_usage_hours"],
            case["maintenance_frequency_days"],
            round(case["wear"], 2),
            previous_failures
        ])

    # Настройка ширины столбцов
    column_widths = [20, 20, 30, 20, 18, 22, 28, 22, 12, 50]  # Примерные ширины для каждого столбца
    for col_num, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + col_num)].width = width

    # Выравнивание текста по центру
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

   # Путь к папке src рядом со скриптом
    folder_path = os.path.join(os.path.dirname(__file__), '../data/')
    os.makedirs(folder_path, exist_ok=True)

    file_path = os.path.join(folder_path, filename)

    # Удаляем старый файл, если он существует
    if os.path.exists(file_path):
        os.remove(file_path)

    # Теперь сохраняем новый файл
    wb.save(file_path)
    print(f"Данные сохранены в файл '{file_path}' ✅")

if __name__ == "__main__":
    num_cases = 600  # Сколько тестовых данных сгенерировать
    cases = [generate_equipment_case() for _ in range(num_cases)]
    
    # Очищаем данные
    cleaned_cases = clean_data(cases)
    
    save_cases_to_excel(cleaned_cases)

    print(f"{len(cleaned_cases)} тестовых кейсов сохранено в Excel-файл 'test_cases.xlsx' ✅")
